# app.py
# ✅ Nom image sans extension
# ✅ Popup (modal) si doublon -> mise à jour ou ignorer
# ✅ Reset complet après action (image + décocher) SANS erreur Streamlit (versioning keys)
# ✅ Excel auto-créé + migré
# ✅ Vidage cache au démarrage

import base64
from pathlib import Path
from datetime import datetime

import streamlit as st
from openpyxl import Workbook, load_workbook


# --------------------------------------------------
# Page config + cache clearing
# --------------------------------------------------
st.set_page_config(page_title="Analyse Clinique", layout="wide")
st.cache_data.clear()
st.cache_resource.clear()

# --------------------------------------------------
# Constants
# --------------------------------------------------
EXCEL_PATH = Path("analyse_clinique.xlsx")
SHEET_NAME = "analyses"

PATHOLOGIES = [
"alveolar pattern",
"apical pleural thickening",
"atelectasis",
"bullas",
"cardiomegaly",
"cavitation",
"consolidation",
"hilar enlargement",
"hydropneumothorax",
"interstitial pattern",
"lobar atelectasis",
"mass",
"mediastinal enlargement",
"mediastinal mass",
"miliary opacities",
"nodule",
"normal",
"pericardial effusion",
"pleural effusion",
"pneumonia",
"pneumoperitone",
"pneumothorax",
"pulmonary edema",
"pulmonary fibrosis",
"reticular interstitial pattern",
"reticulonodular interstitial pattern",
"rib fracture",
"tuberculosis",
"tuberculosis sequelae",
"vascular hilar enlargement",
 "Autres",
]

HEADER = ["image_name", "analysis_date"] + PATHOLOGIES
PATHOLOGIES_BOX_HEIGHT = 830  # ajuste si besoin


# --------------------------------------------------
# Excel helpers
# --------------------------------------------------
def _get_or_create_sheet(wb, name: str):
    return wb[name] if name in wb.sheetnames else wb.create_sheet(title=name)


def ensure_excel_exists_and_migrated() -> None:
    if not EXCEL_PATH.exists():
        wb = Workbook()
        if wb.active and wb.active.title == "Sheet":
            wb.remove(wb.active)
        ws = wb.create_sheet(title=SHEET_NAME)
        ws.append(HEADER)
        wb.save(EXCEL_PATH)
        return

    wb = load_workbook(EXCEL_PATH)
    ws = _get_or_create_sheet(wb, SHEET_NAME)

    if ws.max_row < 1 or ws.cell(1, 1).value is None:
        ws.delete_rows(1, ws.max_row)
        ws.append(HEADER)
        wb.save(EXCEL_PATH)
        return

    existing_header = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    existing_header = [h if h else "" for h in existing_header]
    if existing_header == HEADER:
        return

    old_idx = {h.strip(): i + 1 for i, h in enumerate(existing_header) if isinstance(h, str) and h.strip()}

    tmp_name = f"{SHEET_NAME}_tmp"
    if tmp_name in wb.sheetnames:
        wb.remove(wb[tmp_name])
    ws_tmp = wb.create_sheet(title=tmp_name)
    ws_tmp.append(HEADER)

    for r in range(2, ws.max_row + 1):
        row = []
        for h in HEADER:
            if h in old_idx:
                row.append(ws.cell(r, old_idx[h]).value)
            else:
                row.append("" if h in ("image_name", "analysis_date") else 0)
        ws_tmp.append(row)

    wb.remove(ws)
    ws_tmp.title = SHEET_NAME
    wb.save(EXCEL_PATH)


def find_last_row_by_image_name(image_name: str):
    wb = load_workbook(EXCEL_PATH)
    ws = wb[SHEET_NAME]
    for r in range(ws.max_row, 1, -1):
        if ws.cell(r, 1).value == image_name:
            return r
    return None


def append_row(image_name: str, selected: set[str]) -> None:
    wb = load_workbook(EXCEL_PATH)
    ws = wb[SHEET_NAME]
    row = [image_name, datetime.now().strftime("%Y-%m-%d %H:%M:%S")]
    row += [1 if p in selected else 0 for p in PATHOLOGIES]
    ws.append(row)
    wb.save(EXCEL_PATH)


def update_row(row_idx: int, image_name: str, selected: set[str]) -> None:
    wb = load_workbook(EXCEL_PATH)
    ws = wb[SHEET_NAME]
    ws.cell(row_idx, 1).value = image_name
    ws.cell(row_idx, 2).value = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    for i, p in enumerate(PATHOLOGIES, start=3):
        ws.cell(row_idx, i).value = 1 if p in selected else 0
    wb.save(EXCEL_PATH)


# --------------------------------------------------
# Init state
# --------------------------------------------------
ensure_excel_exists_and_migrated()

st.session_state.setdefault("selected", set())

# versioning keys (IMPORTANT pour reset sans erreur)
st.session_state.setdefault("chk_version", 0)

# reset file_uploader
st.session_state.setdefault("uploader_key", 0)

# pending duplicate modal
st.session_state.setdefault("pending_name", None)
st.session_state.setdefault("pending_row", None)


# --------------------------------------------------
# Reset UI (image + décocher) via versioning
# --------------------------------------------------
def reset_ui():
    st.session_state.selected = set()
    st.session_state.chk_version += 1      # ✅ nouvelles clés -> toutes les cases repartent décochées
    st.session_state.uploader_key += 1     # ✅ vide l'image
    st.session_state.pending_name = None
    st.session_state.pending_row = None


# --------------------------------------------------
# CSS
# --------------------------------------------------
st.markdown(
    """
    <style>
      .img-box { width:100%; margin:0; padding:0; }
      .img-box img { width:100%; height:auto; display:block; }
      div[data-testid="stCheckbox"] label { line-height:1.15; }
    </style>
    """,
    unsafe_allow_html=True,
)


def render_image(uploaded):
    with st.container(border=True):
        if uploaded is None:
            st.caption("Charge une image pour afficher l’aperçu ici.")
            return
        raw = uploaded.getvalue()
        b64 = base64.b64encode(raw).decode("utf-8")
        mime = "image/jpeg" if uploaded.name.lower().endswith((".jpg", ".jpeg")) else "image/png"
        st.markdown(
            f"<div class='img-box'><img src='data:{mime};base64,{b64}' alt='preview'></div>",
            unsafe_allow_html=True,
        )


# --------------------------------------------------
# Modal dialog (st.dialog / st.experimental_dialog)
# --------------------------------------------------
_dialog_decorator = getattr(st, "dialog", None) or getattr(st, "experimental_dialog", None)

if _dialog_decorator is None:
    # fallback (rare)
    def open_update_dialog():
        st.warning(
            f"L’image **{st.session_state.pending_name}** existe déjà. Mettre à jour l’analyse ?"
        )
        c1, c2 = st.columns(2)
        with c1:
            if st.button("Oui, mettre à jour", use_container_width=True):
                update_row(st.session_state.pending_row, st.session_state.pending_name, st.session_state.selected)
                st.success("Analyse mise à jour ✅")
                reset_ui()
                st.rerun()
        with c2:
            if st.button("Non, ignorer", use_container_width=True):
                st.info("Analyse ignorée.")
                reset_ui()
                st.rerun()
else:
    @_dialog_decorator("Image déjà existante")
    def open_update_dialog():
        st.write(
            f"Le nom d’image **{st.session_state.pending_name}** existe déjà.\n\n"
            "Souhaitez-vous **mettre à jour** l’analyse ?"
        )
        c1, c2 = st.columns(2)
        with c1:
            if st.button("Oui, mettre à jour", type="primary", use_container_width=True):
                update_row(st.session_state.pending_row, st.session_state.pending_name, st.session_state.selected)
                reset_ui()
                st.rerun()
        with c2:
            if st.button("Non, ignorer", use_container_width=True):
                reset_ui()
                st.rerun()


# --------------------------------------------------
# UI
# --------------------------------------------------
st.title("Analyse Clinique — Image & Pathologies")

left, right = st.columns([3.2, 1.2], gap="large")

with left:
    st.subheader("Aperçu de l'image")
    uploaded = st.file_uploader(
        "Charger une image",
        type=["png", "jpg", "jpeg"],
        key=f"uploader_{st.session_state.uploader_key}",
    )
    render_image(uploaded)

with right:
    st.subheader("Pathologies")

    a, b = st.columns(2)
    if a.button("Tout cocher", use_container_width=True):
        st.session_state.selected = set(PATHOLOGIES)
        st.session_state.chk_version += 1  # force refresh of checkbox widgets
        st.rerun()

    if b.button("Tout décocher", use_container_width=True):
        st.session_state.selected = set()
        st.session_state.chk_version += 1
        st.rerun()

    with st.container(border=True, height=PATHOLOGIES_BOX_HEIGHT):
        c1, c2 = st.columns(2, gap="small")
        mid = (len(PATHOLOGIES) + 1) // 2

        def checkbox_key(p: str) -> str:
            return f"chk_{p}_{st.session_state.chk_version}"

        for col, items in zip((c1, c2), (PATHOLOGIES[:mid], PATHOLOGIES[mid:])):
            with col:
                for p in items:
                    val = st.checkbox(
                        p,
                        value=(p in st.session_state.selected),
                        key=checkbox_key(p),
                    )
                    if val:
                        st.session_state.selected.add(p)
                    else:
                        st.session_state.selected.discard(p)

    st.write("")
    if st.button("Suivant ➜ Enregistrer", type="primary", use_container_width=True, disabled=(uploaded is None)):
        image_base_name = Path(uploaded.name).stem  # ✅ sans extension
        existing_row = find_last_row_by_image_name(image_base_name)

        if existing_row is None:
            append_row(image_base_name, st.session_state.selected)
            st.success("Analyse enregistrée ✅")
            reset_ui()
            st.rerun()
        else:
            st.session_state.pending_name = image_base_name
            st.session_state.pending_row = existing_row
            open_update_dialog()
